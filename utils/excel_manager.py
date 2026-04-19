import os
import openpyxl
from openpyxl.utils import get_column_letter

class ExcelManager:
    def __init__(self, filepath="data/student_marks.xlsx"):
        self.filepath = filepath
        self.ensure_setup()

    def ensure_setup(self):
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        if not os.path.exists(self.filepath):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            wb.save(self.filepath)

    def append_data(self, image_name, table_data):
        """
        table_data should be a list of lists representing rows and columns.
        """
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active

        if not isinstance(table_data, list) or len(table_data) == 0:
            return

        # Check if the sheet is completely empty
        is_empty = (ws.max_row == 1 and ws.cell(row=1, column=1).value is None)

        # If the sheet already has data, skip the first row (the headers)
        if not is_empty:
            rows_to_append = table_data[1:]
        else:
            rows_to_append = table_data

        for row in rows_to_append:
            if isinstance(row, list):
                ws.append(row)
            elif isinstance(row, dict):
                ws.append(list(row.values()))

        wb.save(self.filepath)

# Singleton instance
excel_manager = ExcelManager()
